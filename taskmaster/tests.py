"""
This module contains test cases for the TaskMaster application.
It includes tests for user registration, login, task management, and exporting tasks to Excel.
"""

from django.test import TestCase, Client, RequestFactory
from django.urls import reverse
import datetime
from parameterized import parameterized
from django.contrib.auth.models import User
from rest_framework.test import APITestCase
from rest_framework.exceptions import ErrorDetail
from openpyxl import load_workbook
from io import BytesIO
from .models import *
from .views import export_task_to_excel

# Create your tests here.


class RegistrationLoginTest(TestCase):
    """
    Test cases for user registration and login functionality.
    """

    def setUp(self):
        """
        Set up the test environment by creating a valid user for testing registration and login.
        """
        self.valid_username = 'testuser1'
        self.valid_password = 'Albert.23'
        user = User.objects.create_user(
            username=self.valid_username, password=self.valid_password)
        user.save()
        UserProfile.objects.create(user=user)

    @parameterized.expand([
        # Test cases for invalid registration
        ('', '', '', '', ['This field is required.', 'This field is required.',
         'This field is required.', 'This field is required.']),
        ('testuser', 'testuser2', '1234', '1234', ['This password is entirely numeric.', 'This password must contain at least 1 symbol: @, #, etc', 'This password is too common.',
         'This password is too short. It must contain at least 8 characters.', 'This password must contain at least one lowercase letter: a-z.', 'This password must contain at least one uppercase letter: A-Z.']),
        ('testuser', 'testuser2', 'Albertt@', 'Albertt@',
         ['This password must contain at least one digit: 0-9.']),
        ('testuser', 'testuser2', 'Albert@23', 'password',
         ['The two password fields didn’t match.']),
        ('testuser1', 'testuser2', 'Albert@23', 'Albert@23',
         ['Full name should only contain alphabetic characters.']),
        ('testuser', 'testuser1', 'Albert@23', 'Albert@23',
         ['A user with that username already exists.']),
    ])
    def test_invalid_registration(self, fullname, username, password1, password2, expected_errors):
        """
        Test invalid registration form submissions.

        Parameters:
        - fullname (str): Full name of the user.
        - username (str): Username for the user.
        - password1 (str): First password field.
        - password2 (str): Second password field.
        - expected_errors (list): Expected error messages.
        """
        data = {
            'fullname': fullname,
            'username': username,
            'password1': password1,
            'password2': password2,
        }
        response = self.client.post(reverse('register'), data)
        self.assertEqual(response.status_code, 200)
        form = response.context['register_form']
        errors = [error for field_errors in form.errors.values()
                  for error in field_errors]
        self.assertEqual(set(errors), set(expected_errors))
        if not username == 'testuser1':
            self.assertFalse(User.objects.filter(username=username).exists())

    def test_valid_registration(self):
        """
        Test valid registration form submission.
        """
        data = {
            'fullname': 'testuser',
            'username': 'testuser2',
            'password1': self.valid_password,
            'password2': self.valid_password,
        }
        response = self.client.post(reverse('register'), data)
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse('login'))
        self.assertTrue(User.objects.filter(
            username=self.valid_username).exists())

    def test_valid_login(self):
        """
        Test valid login form submission.
        """
        data = {'username': self.valid_username,
                'password': self.valid_password}
        response = self.client.post(reverse('login'), data)
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse('index'))
        self.assertTrue(response.wsgi_request.user.is_authenticated)

    @parameterized.expand([
        # Test cases for invalid login
        ('', '', ['This field is required.', 'This field is required.']),
        ('testuser1', 'password', [
         'Please enter a correct username and password. Note that both fields may be case-sensitive.']),
        ('testuser2', 'Albert@23',
         ['Please enter a correct username and password. Note that both fields may be case-sensitive.']),
    ])
    def test_invalid_login(self, username, password, expected_errors):
        """
        Test invalid login form submissions.

        Parameters:
        - username (str): Username for the user.
        - password (str): Password for the user.
        - expected_errors (list): Expected error messages.
        """
        data = {'username': username, 'password': password}
        response = self.client.post(reverse('login'), data)
        form = response.context['login_form']
        errors = [error for field_errors in form.errors.values()
                  for error in field_errors]
        self.assertEqual(set(errors), set(expected_errors))
        self.assertFalse(response.wsgi_request.user.is_authenticated)

    def test_logout(self):
        """
        Test user logout functionality.
        """
        self.client.login(username=self.valid_username,
                          password=self.valid_password)
        response = self.client.post(reverse('logout'))
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse('auth'))
        self.assertFalse(response.wsgi_request.user.is_authenticated)


class AddTask(TestCase):
    def setUp(self):
        """
        Set up the test environment by creating a test user and logging in.
        """
        # Create a test user
        self.user = User.objects.create_user(
            username='testuser', password='testpassword')
        UserProfile.objects.create(user=self.user)
        self.client = Client()
        self.client.login(username='testuser', password='testpassword')

    @parameterized.expand([
        # Test cases with missing fields
        ('', '', '', '', '', '', '', ''),
        (True, '', '', '', '', '', '', ''),
        ('', True, '', '', '', '', '', ''),
        ('', '', True, '', '', '', '', ''),
        # Test cases with invalid data for daily, weekly, and monthly tasks
        (True, '', '', 'title', 'description', 'invalid', 'Sunday', 1),
        ('', True, '', 'title', 'description', 'invalid', 'Sunday', 1),
        ('', '', True, 'title', 'description', 'invalid', 'Sunday', 1),
        ('', True, '', 'title', 'description', '20:20', 'invalid', 1),
        ('', '', True, 'title', 'description', '20:20', 'Sunday', 0),
        (True, '', '', 'title', 'description', '20:20', 'Sunday', 1),
        ('', True, '', 'title', 'description', '20:20', 'Sunday', 1),
        ('', True, '', 'title', 'description', '20:20', 'Sunday', 1),
    ])
    def test_invalid_add_task(self, daily, weekly, monthly, title, description, execution_time, execution_day, execution_date):
        """
        Test the behavior of the 'add_task' view with invalid data.

        Parameters:
        - daily (bool): Whether the task is a daily task.
        - weekly (bool): Whether the task is a weekly task.
        - monthly (bool): Whether the task is a monthly task.
        - title (str): The task title.
        - description (str): The task description.
        - execution_time (str): The task execution time.
        - execution_day (str): The task execution day (only applicable for weekly tasks).
        - execution_date (int): The task execution date (only applicable for monthly tasks).
        """
        # Make a POST request to the add_task view with invalid data
        data = {
            'daily': daily,
            'weekly': weekly,
            'monthly': monthly,
            'title': title,
            'description': description,
            'execution_time': execution_time,
            'execution_day': execution_day,
            'execution_date': execution_date
        }

        response = self.client.post(reverse('add_task'), data)

        # Check if the response status code is 302 (redirect)
        self.assertEqual(response.status_code, 302)

        # Check if the task has not been created
        task = Task.objects.filter(title=title, user=self.user).exists()
        self.assertFalse(task)

        if daily:
            # Check if the response redirects to the 'dailytask' URL
            self.assertRedirects(response, reverse('dailytask'))
        elif weekly:
            # Check if the response redirects to the 'weeklytask' URL
            self.assertRedirects(response, reverse('weeklytask'))
        elif monthly:
            # Check if the response redirects to the 'monthlytask' URL
            self.assertRedirects(response, reverse('monthlytask'))
        else:
            # Check if the response redirects to the 'index' URL
            self.assertRedirects(response, reverse('index'))

    @parameterized.expand([
        # Test cases with valid data for daily, weekly, and monthly tasks
        (True, '', '', 'daily', 'description', '20:20', '', ''),
        ('', True, '', 'weekly', 'description', '20:20', 'Sunday', ''),
        ('', '', True, 'monthly', 'description', '20:20', '', 1),
    ])
    def test_valid_add_task(self, daily, weekly, monthly, title, description, execution_time, execution_day, execution_date):
        """
        Test the behavior of the 'add_task' view with valid data.

        Parameters:
        - daily (bool): Whether the task is a daily task.
        - weekly (bool): Whether the task is a weekly task.
        - monthly (bool): Whether the task is a monthly task.
        - title (str): The task title.
        - description (str): The task description.
        - execution_time (str): The task execution time.
        - execution_day (str): The task execution day (only applicable for weekly tasks).
        - execution_date (int): The task execution date (only applicable for monthly tasks).
        """
        # Make a POST request to the add_task view with valid data
        data = {
            'daily': daily,
            'weekly': weekly,
            'monthly': monthly,
            'title': title,
            'description': description,
            'execution_time': execution_time,
            'execution_day': execution_day,
            'execution_date': execution_date
        }

        response = self.client.post(reverse('add_task'), data)

        # Check if the response status code is 302 (redirect)
        self.assertEqual(response.status_code, 302)

        # Check if the task has not been created
        task = Task.objects.filter(title=title, user=self.user).exists()
        self.assertTrue(task)

        if daily:
            # Check if the response redirects to the 'dailytask' URL
            self.assertRedirects(response, reverse('dailytask'))
        elif weekly:
            # Check if the response redirects to the 'weeklytask' URL
            self.assertRedirects(response, reverse('weeklytask'))
        elif monthly:
            # Check if the response redirects to the 'monthlytask' URL
            self.assertRedirects(response, reverse('monthlytask'))


class EditTask(TestCase):
    def setUp(self):
        """Set up the test environment and create a test user and task."""
        # Create a test user
        self.user = User.objects.create_user(
            username='testuser', password='testpassword')
        UserProfile.objects.create(user=self.user)
        self.title = 'title'
        self.description = 'description'
        self.execution_time = '00:00'
        self.execution_day = 'Monday'
        self.execution_date = 31
        self.client = Client()
        self.client.login(username='testuser', password='testpassword')
        self.task = Task.objects.create(user=self.user, title=self.title, description=self.description,
                                        execution_time=self.execution_time, execution_day=self.execution_day, execution_date=self.execution_date)

    @parameterized.expand([
        # Test missing field including type
        ('', '', '', '', '', '', '', ''),
        # Test missing field for dailytask
        (True, '', '', '', '', '', '', ''),
        # Test missing field for weeklytask
        ('', True, '', '', '', '', '', ''),
        # Test missing field for monthlytask
        ('', '', True, '', '', '', '', ''),
        # Test invalid execution time for daily task
        (True, '', '', 'title', 'description', 'invalid', 'Sunday', 1),
        # Test invalid execution time for weekly task
        ('', True, '', 'title', 'description', 'invalid', 'Sunday', 1),
        # Test invalid execution time for monthly task
        ('', '', True, 'title', 'description', 'invalid', 'Sunday', 1),
        # Test invalid execution day
        ('', True, '', 'title', 'description', '20:20', 'invalid', 1),
        # Test invalid execution date
        ('', '', True, 'title', 'description', '20:20', 'Sunday', 0),
        # Test wrong field in dailytask
        (True, '', '', 'title', 'description', '20:20', 'Sunday', 1),
        # Test wrong field in weeklytask
        ('', True, '', 'title', 'description', '20:20', 'Sunday', 1),
        # Test wrong field in monthlytask
        ('', True, '', 'title', 'description', '20:20', 'Sunday', 1),
    ])
    def test_invalid_edit_task(self, daily, weekly, monthly, title, description, execution_time, execution_day, execution_date):
        """Test editing a task with invalid data."""
        # Make a POST request to the edit_task view with invalid data
        data = {
            'daily': daily,
            'weekly': weekly,
            'monthly': monthly,
            'title': title,
            'description': description,
            'execution_time': execution_time,
            'execution_day': execution_day,
            'execution_date': execution_date
        }

        response = self.client.post(
            reverse('edit_task', args=[self.task.id]), data)

        # Check if the response status code is 302 (redirect)
        self.assertEqual(response.status_code, 302)

        # Refresh the task from the database
        self.task.refresh_from_db()

        # Convert string to date time object
        self.execution_time = datetime.time.fromisoformat(self.execution_time)

        # Check if the task's fields have not updated
        self.assertEqual(self.task.daily, False)
        self.assertEqual(self.task.weekly, False)
        self.assertEqual(self.task.monthly, False)
        self.assertEqual(self.task.title, self.title)
        self.assertEqual(self.task.description, self.description)
        self.assertEqual(self.task.execution_time, self.execution_time)
        self.assertEqual(self.task.execution_day, self.execution_day)
        self.assertEqual(self.task.execution_date, self.execution_date)

        if daily:
            # Check if the response redirects to the 'dailytask' URL
            self.assertRedirects(response, reverse('dailytask'))
        elif weekly:
            # Check if the response redirects to the 'weeklytask' URL
            self.assertRedirects(response, reverse('weeklytask'))
        elif monthly:
            # Check if the response redirects to the 'monthlytask' URL
            self.assertRedirects(response, reverse('monthlytask'))
        else:
            # Check if the response redirects to the 'index' URL
            self.assertRedirects(response, reverse('index'))

    @parameterized.expand([
        # Test valid daily task
        (True, '', '', 'daily', 'description', '20:20', '', ''),
        # Test valid weekly task
        ('', True, '', 'weekly', 'description', '20:20', 'Sunday', ''),
        # Test valid monthly task
        ('', '', True, 'monthly', 'description', '20:20', '', 1),
    ])
    def test_valid_edit_task(self, daily, weekly, monthly, title, description, execution_time, execution_day, execution_date):
        """Test editing a task with valid data."""
        # Make a POST request to the edit_task view with valid data
        data = {
            'daily': daily,
            'weekly': weekly,
            'monthly': monthly,
            'title': title,
            'description': description,
            'execution_time': execution_time,
            'execution_day': execution_day,
            'execution_date': execution_date
        }

        response = self.client.post(
            reverse('edit_task', args=[self.task.id]), data)

        # Check if the response status code is 302 (redirect)
        self.assertEqual(response.status_code, 302)

        # Refresh the task from the database
        self.task.refresh_from_db()

        # Convert string to date time object
        execution_time = datetime.time.fromisoformat(execution_time)

        # Check if the task's fields have been updated
        self.assertEqual(self.task.daily, True if daily else False)
        self.assertEqual(self.task.weekly, True if weekly else False)
        self.assertEqual(self.task.monthly, True if monthly else False)
        self.assertEqual(self.task.title, title)
        self.assertEqual(self.task.description, description)
        self.assertEqual(self.task.execution_time, execution_time)
        self.assertEqual(self.task.execution_day,
                         execution_day if weekly else None)
        self.assertEqual(self.task.execution_date,
                         execution_date if monthly else None)

        if daily:
            # Check if the response redirects to the 'dailytask' URL
            self.assertRedirects(response, reverse('dailytask'))
        elif weekly:
            # Check if the response redirects to the 'weeklytask' URL
            self.assertRedirects(response, reverse('weeklytask'))
        elif monthly:
            # Check if the response redirects to the 'monthlytask' URL
            self.assertRedirects(response, reverse('monthlytask'))


class DeleteTask(TestCase):
    def setUp(self):
        """
        Set up the test environment with necessary data and log in as a test user.
        """
        # Create a test user
        self.user = User.objects.create_user(
            username='testuser', password='testpassword')
        UserProfile.objects.create(user=self.user)

        # Initialize the client and log in as the test user
        self.client = Client()
        self.client.login(username='testuser', password='testpassword')

        # Create a test task associated with the test user
        self.task = Task.objects.create(
            user=self.user, daily=True, title='title', description='description', execution_time='00:00')

    def test_invalid_delete_task_not_found(self):
        """
        Test deleting a task that does not exist in the database (should return a 404 error).
        """
        # Attempt to delete a task that does not exist in the database
        invalid_task_id = 9999

        # Make a POST request to the delete_task view with an invalid task ID
        response = self.client.post(
            reverse('delete_task', args=[invalid_task_id]))

        # Check if the response status code is 404 (not found)
        self.assertEqual(response.status_code, 404)

        # Check if the task still exists in the database
        task = Task.objects.filter(id=self.task.id).exists()
        self.assertTrue(task)

    def test_invalid_delete_task_permission_denied(self):
        """
        Test deleting a task with a different user's authentication (should return a 404 error).
        """
        # Create a different user for testing (no permission to delete the task)
        other_user = User.objects.create_user(
            username='otheruser', password='otherpassword')
        UserProfile.objects.create(user=other_user)

        # Simulate a login request for the different user
        self.client.login(username='otheruser', password='otherpassword')

        # Make a POST request to the delete_task view with the task's ID using the different user's authentication
        response = self.client.post(
            reverse('delete_task', args=[self.task.id]))

        # Check if the response status code is 404 (not found)
        self.assertEqual(response.status_code, 404)

        # Check if the task still exists in the database
        task = Task.objects.filter(id=self.task.id).exists()
        self.assertTrue(task)

    def test_valid_delete_task(self):
        """
        Test deleting a valid task from the database (should be successful).
        """
        # Make a POST request to the delete_task view with the task's ID
        response = self.client.post(
            reverse('delete_task', args=[self.task.id]))

        # Check if the response status code is 302 (redirect)
        self.assertEqual(response.status_code, 302)

        # Check if the task has been deleted from the database
        task = Task.objects.filter(id=self.task.id).exists()
        self.assertFalse(task)


class MarkCompleteTask(TestCase):
    def setUp(self):
        """Set up the test environment with a test user and a task."""
        # Create a test user
        self.user = User.objects.create_user(
            username='testuser', password='testpassword')
        UserProfile.objects.create(user=self.user)
        self.client = Client()
        self.client.login(username='testuser', password='testpassword')
        self.task = Task.objects.create(
            user=self.user, daily=True, title='title', description='description', execution_time='00:00')

    def test_invalid_mark_complete_task_not_found(self):
        """Test marking a non-existing task as complete."""
        # Attempt to delete a task that does not exist in the database
        invalid_task_id = 9999

        # Make a POST request to the mark_task_complete view with an invalid task ID
        response = self.client.post(
            reverse('mark_task_complete', args=[invalid_task_id]))

        # Check if the response status code is 404 (not found)
        self.assertEqual(response.status_code, 404)

        # Refresh the task from the database
        self.task.refresh_from_db()

        # Check if the task not completed
        task = Task.objects.get(id=self.task.id)
        self.assertFalse(task.completed)

    def test_invalid_delete_task_permission_denied(self):
        """Test marking a task as complete with a different user's authentication (no permission)."""
        # Create a different user for testing (no permission to delete the task)
        other_user = User.objects.create_user(
            username='otheruser', password='otherpassword')
        UserProfile.objects.create(user=other_user)

        # Simulate a login request for the different user
        self.client.login(username='otheruser', password='otherpassword')

        # Make a POST request to the mark_task_complete view with the task's ID using the different user's authentication
        response = self.client.post(
            reverse('mark_task_complete', args=[self.task.id]))

        # Check if the response status code is 404 (not found)
        self.assertEqual(response.status_code, 404)

        # Refresh the task from the database
        self.task.refresh_from_db()

        # Check if the task not completed
        task = Task.objects.get(id=self.task.id)
        self.assertFalse(task.completed)

    def test_valid_delete_task(self):
        """Test marking a task as complete."""
        # Make a POST request to the mark_task_complete view with the task's ID
        response = self.client.post(
            reverse('mark_task_complete', args=[self.task.id]))

        # Check if the response status code is 302 (redirect)
        self.assertEqual(response.status_code, 302)

        # Refresh the task from the database
        self.task.refresh_from_db()

        # Check if the task is completed
        task = Task.objects.get(id=self.task.id)
        self.assertTrue(task.completed)


class TaskViewSetTestCase(APITestCase):
    def setUp(self):
        # Set up a test user and task before each test
        self.user = User.objects.create_user(
            username='testuser', password='testpassword')
        self.title = 'title 2'
        self.description = 'description'
        self.execution_time = '00:00'
        self.execution_day = 'Monday'
        self.execution_date = 31
        self.client = Client()
        self.client.login(username='testuser', password='testpassword')
        self.task = Task.objects.create(user=self.user, title=self.title, description=self.description,
                                        execution_time=self.execution_time, execution_day=self.execution_day, execution_date=self.execution_date)

    @parameterized.expand([
        # Test missing all field
        ('', '', '', '', '', '', '', '', '', {
            "title": [
                ErrorDetail(
                    string="This field may not be blank.", code='blank')
            ],
            "description": [
                ErrorDetail(
                    string="This field may not be blank.", code='blank')
            ],
            "completed": [
                ErrorDetail(string="Must be a valid boolean.", code='invalid')
            ],
            "daily": [
                ErrorDetail(string="Must be a valid boolean.", code='invalid')
            ],
            "weekly": [
                ErrorDetail(string="Must be a valid boolean.", code='invalid')
            ],
            "monthly": [
                ErrorDetail(string="Must be a valid boolean.", code='invalid')
            ],
            "execution_time": [
                ErrorDetail(
                    string="Time has wrong format. Use one of these formats instead: hh:mm[:ss[.uuuuuu]].", code='invalid')
            ],
            "execution_date": [
                ErrorDetail(string="A valid integer is required.",
                            code='invalid')
            ],
        }),
        # Test missing field for dailytask
        ('title', 'description', False, True, False, False, None, None, None, {
            "execution_time": [
                ErrorDetail(string='This field may not be blank.',
                            code='invalid')
            ],
        }),
        # Test missing field for weeklytask
        ('title', 'description', False, False, True, False, None, '20:20', None, {
            "execution_day": [
                ErrorDetail(string='This field may not be blank.',
                            code='invalid')
            ],
        }),
        # Test missing field for monthlytask
        ('title', 'description', False, False, False, True, None, '20:20', None, {
            "execution_date": [
                ErrorDetail(string='This field may not be blank.',
                            code='invalid')
            ],
        }),
        # Test invalid execution time for daily task
        ('title', 'description', False, True, False, False, None, 20, None, {
            "execution_time": [
                ErrorDetail(
                    string='Time has wrong format. Use one of these formats instead: hh:mm[:ss[.uuuuuu]].', code='invalid')
            ],
        }),
        # Test invalid execution time for weekly task
        ('title', 'description', False, False, True, False, 'Monday', 20, None, {
            "execution_time": [
                ErrorDetail(
                    string='Time has wrong format. Use one of these formats instead: hh:mm[:ss[.uuuuuu]].', code='invalid')
            ],
        }),
        # Test invalid execution time for monthly task
        ('title', 'description', False, False, False, True, None, 20, 31, {
            "execution_time": [
                ErrorDetail(
                    string='Time has wrong format. Use one of these formats instead: hh:mm[:ss[.uuuuuu]].', code='invalid')
            ],
        }),
        # Test invalid execution day
        ('title', 'description', False, False, True, False, 'invalid', '20:20', None, {
            "execution_day": [
                ErrorDetail(
                    string='Invalid value. Choose from: Sunday, Monday, Tuesday, Wednesday, Thursday, Friday, Saturday.', code='invalid')
            ],
        }),
        # Test invalid execution date
        ('title', 'description', False, False, False, True, None, '20:20', 'invalid', {
            "execution_date": [
                ErrorDetail(string='A valid integer is required.',
                            code='invalid')
            ],
        }),
        # Test invalid execution date
        ('title', 'description', False, False, False, True, None, '20:20', 32, {
            "execution_date": [
                ErrorDetail(
                    string='Ensure this value is less than or equal to 31.', code='max_value')
            ],
        }),
        # Test invalid execution date
        ('title', 'description', False, False, False, True, None, '20:20', 0, {
            "execution_date": [
                ErrorDetail(
                    string='Ensure this value is greater than or equal to 1.', code='min_value')
            ],
        }),
        # Test wrong field in dailytask
        ('title', 'description', False, True, False, False, 'Monday', '20:20', 31, {
            "dailytask": [
                ErrorDetail(string='Invalid field for daily task.',
                            code='invalid')
            ],
        }),
        # Test wrong field in weeklytask
        ('title', 'description', False, False, True, False, 'Monday', '20:20', 31, {
            "weeklytask": [
                ErrorDetail(
                    string='Invalid field for weekly task.', code='invalid')
            ],
        }),
        # Test wrong field in monthlytask
        ('title', 'description', False, False, False, True, 'Monday', '20:20', 31, {
            "monthlytask": [
                ErrorDetail(
                    string='Invalid field for monthly task.', code='invalid')
            ],
        }),
        # Test invalid type
        ('title', 'description', False, True, True, True, 'Monday', '20:20', 31, {
            "type error": [
                ErrorDetail(
                    string="Please select one and only one of 'Daily', 'Weekly', or 'Monthly'.", code='invalid')
            ],
        }),
        # Test invalid type
        ('title', 'description', False, False, False, False, 'Monday', '20:20', 31, {
            "type error": [
                ErrorDetail(
                    string="Please select one and only one of 'Daily', 'Weekly', or 'Monthly'.", code='invalid')
            ],
        }),
    ])
    def test_invalid_create_task(self, title, description, completed, daily, weekly, monthly, execution_day, execution_time, execution_date, expected_error):
        """
        Test cases for invalid task creation (POST request).

        Parameters:
        - title (str): The task title.
        - description (str): The task description.
        - completed (bool): Task completion status.
        - daily (bool): Whether the task is a daily task.
        - weekly (bool): Whether the task is a weekly task.
        - monthly (bool): Whether the task is a monthly task.
        - execution_day (str): The task execution day (only applicable for weekly tasks).
        - execution_time (str): The task execution time.
        - execution_date (int): The task execution date (only applicable for monthly tasks).
        - expected_error (dict): Expected error details.
        """
        # Test POST request to create a new task
        url = '/api/tasks/'
        data = {
            'title': title,
            'description': description,
            'completed': completed,
            'daily': daily,
            'weekly': weekly,
            'monthly': monthly,
            'execution_day': execution_day,
            'execution_time': execution_time,
            'execution_date': execution_date,
            'user': self.user.id
        }
        response = self.client.post(
            url, data, format='json', content_type='application/json')
        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.data, expected_error)

        # Check if the task has not been created
        task = Task.objects.filter(title=title, user=self.user).exists()
        self.assertFalse(task)

    @parameterized.expand([
        # Test valid dailytask
        ('title', 'description', False, True, False, False, None, '20:20', None),
        # Test valid weeklytask
        ('title', 'description', False, False,
         True, False, 'Monday', '20:20', None),
        # Test valid monthlytask
        ('title', 'description', False, False, False, True, None, '20:20', 31),
    ])
    def test_valid_create_task(self, title, description, completed, daily, weekly, monthly, execution_day, execution_time, execution_date):
        """
        Test cases for valid task creation (POST request).

        Parameters:
        - title (str): The task title.
        - description (str): The task description.
        - completed (bool): Task completion status.
        - daily (bool): Whether the task is a daily task.
        - weekly (bool): Whether the task is a weekly task.
        - monthly (bool): Whether the task is a monthly task.
        - execution_day (str): The task execution day (only applicable for weekly tasks).
        - execution_time (str): The task execution time.
        - execution_date (int): The task execution date (only applicable for monthly tasks).
        """
        # Test POST request to create a new task
        url = '/api/tasks/'
        data = {
            'title': title,
            'description': description,
            'completed': completed,
            'daily': daily,
            'weekly': weekly,
            'monthly': monthly,
            'execution_day': execution_day,
            'execution_time': execution_time,
            'execution_date': execution_date,
            'user': self.user.id
        }
        response = self.client.post(url, data, format='json')
        self.assertEqual(response.status_code, 201)

        # Check if the task has not been created
        task = Task.objects.filter(title=title, user=self.user).exists()
        self.assertTrue(task)

    @parameterized.expand([
        # Test missing all field
        ('', '', '', '', '', '', '', '', '', {
            "title": [
                ErrorDetail(
                    string="This field may not be blank.", code='blank')
            ],
            "description": [
                ErrorDetail(
                    string="This field may not be blank.", code='blank')
            ],
            "completed": [
                ErrorDetail(string="Must be a valid boolean.", code='invalid')
            ],
            "daily": [
                ErrorDetail(string="Must be a valid boolean.", code='invalid')
            ],
            "weekly": [
                ErrorDetail(string="Must be a valid boolean.", code='invalid')
            ],
            "monthly": [
                ErrorDetail(string="Must be a valid boolean.", code='invalid')
            ],
            "execution_time": [
                ErrorDetail(
                    string="Time has wrong format. Use one of these formats instead: hh:mm[:ss[.uuuuuu]].", code='invalid')
            ],
            "execution_date": [
                ErrorDetail(string="A valid integer is required.",
                            code='invalid')
            ],
        }),
        # Test missing field for dailytask
        ('title', 'description', False, True, False, False, None, None, None, {
            "execution_time": [
                ErrorDetail(string='This field may not be blank.',
                            code='invalid')
            ],
        }),
        # Test missing field for weeklytask
        ('title', 'description', False, False, True, False, None, '20:20', None, {
            "execution_day": [
                ErrorDetail(string='This field may not be blank.',
                            code='invalid')
            ],
        }),
        # Test missing field for monthlytask
        ('title', 'description', False, False, False, True, None, '20:20', None, {
            "execution_date": [
                ErrorDetail(string='This field may not be blank.',
                            code='invalid')
            ],
        }),
        # Test invalid execution time for daily task
        ('title', 'description', False, True, False, False, None, 20, None, {
            "execution_time": [
                ErrorDetail(
                    string='Time has wrong format. Use one of these formats instead: hh:mm[:ss[.uuuuuu]].', code='invalid')
            ],
        }),
        # Test invalid execution time for weekly task
        ('title', 'description', False, False, True, False, 'Monday', 20, None, {
            "execution_time": [
                ErrorDetail(
                    string='Time has wrong format. Use one of these formats instead: hh:mm[:ss[.uuuuuu]].', code='invalid')
            ],
        }),
        # Test invalid execution time for monthly task
        ('title', 'description', False, False, False, True, None, 20, 31, {
            "execution_time": [
                ErrorDetail(
                    string='Time has wrong format. Use one of these formats instead: hh:mm[:ss[.uuuuuu]].', code='invalid')
            ],
        }),
        # Test invalid execution day
        ('title', 'description', False, False, True, False, 'invalid', '20:20', None, {
            "execution_day": [
                ErrorDetail(
                    string='Invalid value. Choose from: Sunday, Monday, Tuesday, Wednesday, Thursday, Friday, Saturday.', code='invalid')
            ],
        }),
        # Test invalid execution date
        ('title', 'description', False, False, False, True, None, '20:20', 'invalid', {
            "execution_date": [
                ErrorDetail(string='A valid integer is required.',
                            code='invalid')
            ],
        }),
        # Test invalid execution date
        ('title', 'description', False, False, False, True, None, '20:20', 32, {
            "execution_date": [
                ErrorDetail(
                    string='Ensure this value is less than or equal to 31.', code='max_value')
            ],
        }),
        # Test invalid execution date
        ('title', 'description', False, False, False, True, None, '20:20', 0, {
            "execution_date": [
                ErrorDetail(
                    string='Ensure this value is greater than or equal to 1.', code='min_value')
            ],
        }),
        # Test wrong field in dailytask
        ('title', 'description', False, True, False, False, 'Monday', '20:20', 31, {
            "dailytask": [
                ErrorDetail(string='Invalid field for daily task.',
                            code='invalid')
            ],
        }),
        # Test wrong field in weeklytask
        ('title', 'description', False, False, True, False, 'Monday', '20:20', 31, {
            "weeklytask": [
                ErrorDetail(
                    string='Invalid field for weekly task.', code='invalid')
            ],
        }),
        # Test wrong field in monthlytask
        ('title', 'description', False, False, False, True, 'Monday', '20:20', 31, {
            "monthlytask": [
                ErrorDetail(
                    string='Invalid field for monthly task.', code='invalid')
            ],
        }),
        # Test invalid type
        ('title', 'description', False, True, True, True, 'Monday', '20:20', 31, {
            "type error": [
                ErrorDetail(
                    string="Please select one and only one of 'Daily', 'Weekly', or 'Monthly'.", code='invalid')
            ],
        }),
        # Test invalid type
        ('title', 'description', False, False, False, False, 'Monday', '20:20', 31, {
            "type error": [
                ErrorDetail(
                    string="Please select one and only one of 'Daily', 'Weekly', or 'Monthly'.", code='invalid')
            ],
        }),
    ])
    def test_invalid_edit_task(self, title, description, completed, daily, weekly, monthly, execution_day, execution_time, execution_date, expected_error):
        """
        Test cases for invalid task update (PUT request).

        Parameters:
        - title (str): The task title.
        - description (str): The task description.
        - completed (bool): Task completion status.
        - daily (bool): Whether the task is a daily task.
        - weekly (bool): Whether the task is a weekly task.
        - monthly (bool): Whether the task is a monthly task.
        - execution_day (str): The task execution day (only applicable for weekly tasks).
        - execution_time (str): The task execution time.
        - execution_date (int): The task execution date (only applicable for monthly tasks).
        - expected_error (dict): Expected error details.
        """
        # Test PUT request to create a new task
        url = f'/api/tasks/{self.task.id}/'
        data = {
            'title': title,
            'description': description,
            'completed': completed,
            'daily': daily,
            'weekly': weekly,
            'monthly': monthly,
            'execution_day': execution_day,
            'execution_time': execution_time,
            'execution_date': execution_date,
            'user': self.user.id
        }
        response = self.client.put(
            url, data, format='json', content_type='application/json')
        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.data, expected_error)

        # Refresh the task from the database
        self.task.refresh_from_db()

        # Convert string to date time object
        self.execution_time = datetime.time.fromisoformat(self.execution_time)

        # Check if the task's fields have not updated
        self.assertEqual(self.task.daily, False)
        self.assertEqual(self.task.weekly, False)
        self.assertEqual(self.task.monthly, False)
        self.assertEqual(self.task.title, self.title)
        self.assertEqual(self.task.description, self.description)
        self.assertEqual(self.task.execution_time, self.execution_time)
        self.assertEqual(self.task.execution_day, self.execution_day)
        self.assertEqual(self.task.execution_date, self.execution_date)

    @parameterized.expand([
        # Test valid dailytask
        ('title', 'description', False, True, False, False, None, '20:20', None),
        # Test valid weeklytask
        ('title', 'description', False, False,
         True, False, 'Monday', '20:20', None),
        # Test valid monthlytask
        ('title', 'description', False, False, False, True, None, '20:20', 31),
    ])
    def test_valid_create_task(self, title, description, completed, daily, weekly, monthly, execution_day, execution_time, execution_date):
        """
        Test cases for valid task update (PUT request).

        Parameters:
        - title (str): The task title.
        - description (str): The task description.
        - completed (bool): Task completion status.
        - daily (bool): Whether the task is a daily task.
        - weekly (bool): Whether the task is a weekly task.
        - monthly (bool): Whether the task is a monthly task.
        - execution_day (str): The task execution day (only applicable for weekly tasks).
        - execution_time (str): The task execution time.
        - execution_date (int): The task execution date (only applicable for monthly tasks).
        """
        # Test POST request to create a new task
        url = f'/api/tasks/{self.task.id}/'
        data = {
            'title': title,
            'description': description,
            'completed': completed,
            'daily': daily,
            'weekly': weekly,
            'monthly': monthly,
            'execution_day': execution_day,
            'execution_time': execution_time,
            'execution_date': execution_date,
            'user': self.user.id
        }
        response = self.client.put(
            url, data, format='json', content_type='application/json')
        self.assertEqual(response.status_code, 200)

        # Refresh the task from the database
        self.task.refresh_from_db()

        # Convert string to date time object
        execution_time = datetime.time.fromisoformat(execution_time)

        # Check if the task's fields have been updated
        self.assertEqual(self.task.daily, True if daily else False)
        self.assertEqual(self.task.weekly, True if weekly else False)
        self.assertEqual(self.task.monthly, True if monthly else False)
        self.assertEqual(self.task.title, title)
        self.assertEqual(self.task.description, description)
        self.assertEqual(self.task.execution_time, execution_time)
        self.assertEqual(self.task.execution_day,
                         execution_day if weekly else None)
        self.assertEqual(self.task.execution_date,
                         execution_date if monthly else None)

    def test_valid_delete_task(self):
        """
        Test case for valid task deletion (DELETE request).
        """
        # Make a DELETE request to the api with the task's ID
        url = f'/api/tasks/{self.task.id}/'
        response = self.client.delete(url)

        # Check if the response status code is 204 (deleted)
        self.assertEqual(response.status_code, 204)

        # Check if the task has been deleted from the database
        task = Task.objects.filter(id=self.task.id).exists()
        self.assertFalse(task)

    def test_valid_mark_complete_task(self):
        """
        Test case for marking a task as complete (POST request).
        """
        # Make a POST request to the api with the task's ID
        url = f'/api/tasks/{self.task.id}/complete/'
        response = self.client.post(url)

        # Check if the response status code is 200 (success)
        self.assertEqual(response.status_code, 200)

        # Refresh the task from the database
        self.task.refresh_from_db()

        # Check if the task not completed
        task = Task.objects.get(id=self.task.id)
        self.assertTrue(task.completed)


class ExportTaskToExcelTestCase(TestCase):
    def setUp(self):
        self.factory = RequestFactory()
        # Create a test user
        self.user = User.objects.create_user(
            username='testuser', password='testpassword')

    def test_export_task_to_excel(self):
        """
        Test exporting tasks to Excel.

        Simulates an authenticated user request and creates some test TaskHistory instances for the user.
        It then checks the exported Excel file's content against the expected values.
        """
        # Simulate an authenticated user request
        request = self.factory.get('/export/')
        request.user = self.user

        # Create some test TaskHistory instances for the user
        daily_task = TaskHistory.objects.create(
            user=self.user,
            task_type='daily',
            date='2023-07-05',
            title='Daily Task 1',
            description='This is a daily task',
            execution_time='08:00',
            completed=True,
        )

        weekly_task = TaskHistory.objects.create(
            user=self.user,
            task_type='weekly',
            date='2023-07-05',
            title='Weekly Task 1',
            description='This is a weekly task',
            execution_time='08:00',
            execution_day='Wednesday',
            completed=True,
        )

        monthly_task = TaskHistory.objects.create(
            user=self.user,
            task_type='monthly',
            date='2023-07-05',
            title='Monthly Task 1',
            description='This is a monthly task',
            execution_time='08:00',
            execution_date=31,
            completed=True,
        )

        # Export tasks to Excel
        response = export_task_to_excel(request)

        # Assert the response status code and content type
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        # Load the generated Excel file and check its contents
        excel_file = BytesIO(response.content)
        workbook = load_workbook(excel_file)

        # Assert the sheet names
        # Assuming three task types: daily, weekly, monthly
        self.assertEqual(len(workbook.sheetnames), 3)

        # Assert the content of the 'Daily' sheet
        daily_sheet = workbook['Daily']
        self.assertEqual(daily_sheet.cell(row=1, column=1).value, 'Daily')
        self.assertEqual(daily_sheet.cell(row=2, column=1).value, 'Date')
        self.assertEqual(daily_sheet.cell(row=3, column=1).value, '05-07-2023')
        self.assertEqual(daily_sheet.cell(row=2, column=2).value, 'Title')
        self.assertEqual(daily_sheet.cell(
            row=3, column=2).value, 'Daily Task 1')
        self.assertEqual(daily_sheet.cell(
            row=2, column=3).value, 'Description')
        self.assertEqual(daily_sheet.cell(
            row=3, column=3).value, 'This is a daily task')
        self.assertEqual(daily_sheet.cell(
            row=2, column=4).value, 'Execution_time')
        self.assertEqual(daily_sheet.cell(row=3, column=4).value, '08:00')
        self.assertEqual(daily_sheet.cell(row=2, column=5).value, 'Completed')
        self.assertEqual(daily_sheet.cell(row=3, column=5).value, 'yes')

        # Assert the content of the 'Weekly' sheet
        weekly_sheet = workbook['Weekly']
        self.assertEqual(weekly_sheet.cell(row=1, column=1).value, 'Weekly')
        self.assertEqual(weekly_sheet.cell(row=2, column=1).value, 'Date')
        self.assertEqual(weekly_sheet.cell(
            row=3, column=1).value, '05-07-2023')
        self.assertEqual(weekly_sheet.cell(row=2, column=2).value, 'Title')
        self.assertEqual(weekly_sheet.cell(
            row=3, column=2).value, 'Weekly Task 1')
        self.assertEqual(weekly_sheet.cell(
            row=2, column=3).value, 'Description')
        self.assertEqual(weekly_sheet.cell(
            row=3, column=3).value, 'This is a weekly task')
        self.assertEqual(weekly_sheet.cell(
            row=2, column=4).value, 'Execution_time')
        self.assertEqual(weekly_sheet.cell(
            row=3, column=4).value, 'Wednesday, 08:00')
        self.assertEqual(weekly_sheet.cell(row=2, column=5).value, 'Completed')
        self.assertEqual(weekly_sheet.cell(row=3, column=5).value, 'yes')

        # Assert the content of the 'Monthly' sheet
        monthly_sheet = workbook['Monthly']
        self.assertEqual(monthly_sheet.cell(row=1, column=1).value, 'Monthly')
        self.assertEqual(monthly_sheet.cell(row=2, column=1).value, 'Date')
        self.assertEqual(monthly_sheet.cell(
            row=3, column=1).value, '05-07-2023')
        self.assertEqual(monthly_sheet.cell(row=2, column=2).value, 'Title')
        self.assertEqual(monthly_sheet.cell(
            row=3, column=2).value, 'Monthly Task 1')
        self.assertEqual(monthly_sheet.cell(
            row=2, column=3).value, 'Description')
        self.assertEqual(monthly_sheet.cell(
            row=3, column=3).value, 'This is a monthly task')
        self.assertEqual(monthly_sheet.cell(
            row=2, column=4).value, 'Execution_time')
        self.assertEqual(monthly_sheet.cell(
            row=3, column=4).value, 'Day 31, 08:00')
        self.assertEqual(monthly_sheet.cell(
            row=2, column=5).value, 'Completed')
        self.assertEqual(monthly_sheet.cell(row=3, column=5).value, 'yes')
