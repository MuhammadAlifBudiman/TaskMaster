"""
This module contains the views for the TaskMaster application.

Each view function is responsible for handling specific HTTP requests and rendering appropriate responses.
The views include functionality for user authentication, task management, and exporting task data.
"""

# Import necessary modules and functions from Django
# For rendering templates, redirecting, and fetching objects
from django.shortcuts import render, redirect, get_object_or_404
# For user authentication and session management
from django.contrib.auth import authenticate, login, logout
# For interacting with the User model
from django.contrib.auth.models import User
from django.contrib import messages  # For displaying messages to the user
# For restricting access to authenticated users
from django.contrib.auth.decorators import login_required

# Import necessary modules from openpyxl for Excel file generation
from openpyxl import Workbook  # For creating Excel workbooks
# For styling Excel cells
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
# For working with Excel column letters
from openpyxl.utils import get_column_letter

# Import HttpResponse for sending HTTP responses
from django.http import HttpResponse

# Import forms, models, and serializers from the application
from .forms.forms import *  # Import all forms from the forms module
from .models import *  # Import all models from the models module
# Import all serializers from the serializers module
from .serializers.serializers import *

# Define the index view function


def index(request):
    """
    View function for the task management dashboard.

    Displays statistics for completed and remaining daily, weekly, and monthly tasks,
    and checks if all tasks have been completed.

    :param request: HttpRequest object
    :return: Rendered template response
    """

    # Redirect to the authentication page if the user is not logged in
    if not request.user.is_authenticated:
        return redirect('auth')

    # Get the count of completed and total daily tasks for the user
    completed_daily_tasks = Task.objects.filter(
        completed=True, daily=True, user=request.user).count()
    total_daily_tasks = Task.objects.filter(
        daily=True, user=request.user).count()
    remaining_daily_tasks = total_daily_tasks - completed_daily_tasks

    # Get the count of completed and total weekly tasks for the user
    completed_weekly_tasks = Task.objects.filter(
        completed=True, weekly=True, user=request.user).count()
    total_weekly_tasks = Task.objects.filter(
        weekly=True, user=request.user).count()
    remaining_weekly_tasks = total_weekly_tasks - completed_weekly_tasks

    # Get the count of completed and total monthly tasks for the user
    completed_monthly_tasks = Task.objects.filter(
        completed=True, monthly=True, user=request.user).count()
    total_monthly_tasks = Task.objects.filter(
        monthly=True, user=request.user).count()
    remaining_monthly_tasks = total_monthly_tasks - completed_monthly_tasks

    # Check if all tasks have been completed and there are tasks to count
    completed_all_tasks = (completed_daily_tasks + completed_weekly_tasks + completed_monthly_tasks) == (total_daily_tasks +
                                                                                                         total_weekly_tasks + total_monthly_tasks) and (total_daily_tasks + total_weekly_tasks + total_monthly_tasks) != 0

    # Pass the task statistics to the template for rendering
    return render(request, 'taskmaster/index.html',
                  {
                      'completed_daily_tasks': completed_daily_tasks,
                      'total_daily_tasks': total_daily_tasks,
                      'remaining_daily_tasks': remaining_daily_tasks,
                      'completed_weekly_tasks': completed_weekly_tasks,
                      'total_weekly_tasks': total_weekly_tasks,
                      'remaining_weekly_tasks': remaining_weekly_tasks,
                      'completed_monthly_tasks': completed_monthly_tasks,
                      'total_monthly_tasks': total_monthly_tasks,
                      'remaining_monthly_tasks': remaining_monthly_tasks,
                      'completed_all_tasks': completed_all_tasks,
                  })

# Define the auth_view function


def auth_view(request):
    """
    View function for handling login and registration page.

    If the user is authenticated, it redirects to the index page.
    If the user is not authenticated, it displays the login and registration forms.

    Parameters:
    request (HttpRequest): The HTTP request object.

    Returns:
    HttpResponse: A response containing the rendered login and registration forms.
    """

    # Check if the user is already authenticated
    if request.user.is_authenticated:
        return redirect('index')

    # Create instances of login and registration forms
    login_form = UserLoginForm()
    register_form = UserRegisterForm()

    # Render the auth.html template with the login and registration forms
    return render(request, 'taskmaster/auth.html', {'login_form': login_form, 'register_form': register_form})

# Define the login_view function


def login_view(request):
    """View function to handle user login and registration.

    If the user is already authenticated, redirect to the index page.
    If the request method is POST, process the login form.
    If the login form is valid, log in the user and redirect to the index page.
    If the login form is not valid, display appropriate error messages.
    If the request method is GET, display the login and registration forms.

    Args:
        request (HttpRequest): The HTTP request object.

    Returns:
        HttpResponse: The rendered HTML response or a redirection response.
    """
    if request.user.is_authenticated:
        # If the user is already authenticated, redirect to the index page.
        return redirect('index')

    if request.method == 'POST':
        # If the request method is POST, process the login form.
        login_form = UserLoginForm(data=request.POST)
        register_form = UserRegisterForm()

        if login_form.is_valid():
            # If the login form is valid, log in the user and redirect to the index page.
            user = login_form.get_user()
            login(request, user)
            messages.success(request, f'Logged in successfully.')
            return redirect('index')
        elif not login_form.has_error('username') and not login_form.has_error('password'):
            # If there are no username or password errors but the form is invalid, display an error message.
            messages.error(request, f'Invalid username or password.')
        # Render the login and registration forms with the appropriate messages.
        return render(request, 'taskmaster/auth.html', {'login_form': login_form, 'register_form': register_form})
    else:
        # If the request method is GET, display the login and registration forms.
        login_form = UserLoginForm()
        register_form = UserRegisterForm()

    return render(request, 'taskmaster/auth.html', {'login_form': login_form, 'register_form': register_form})


def register_view(request):
    """
    View function for user registration.

    If the user is already authenticated, they will be redirected to the index page.
    If the request method is POST, the registration form is processed and a new user is registered.
    The user is then redirected to the login page.

    Args:
        request (HttpRequest): The request object.

    Returns:
        HttpResponse: A rendered HTML page with the registration form.
    """

    # If the user is already authenticated, redirect to the index page
    if request.user.is_authenticated:
        return redirect('index')

    checked = True

    if request.method == 'POST':
        # Create instances of the login and registration forms
        login_form = UserLoginForm()
        register_form = UserRegisterForm(request.POST)

        # Check if the registration form is valid
        if register_form.is_valid():
            # Save the registered user
            register_form.save()
            username = register_form.cleaned_data.get('username')

            # Create a user profile for the registered user
            user = User.objects.get(username=username)
            userprofile = UserProfile.objects.create(user=user)
            userprofile.save()

            # Display success message and redirect to the login page
            messages.success(
                request, f'Account created for <b>{username}</b>. You can now log in.')
            return redirect('login')

        elif register_form.has_error('password2'):
            # Add 'is-invalid' class to password1 field if password2 has an error
            register_form.fields['password1'].widget.attrs['class'] = 'is-invalid'
        else:
            # Display error message if an error occurred during registration
            messages.error(
                request, f'An error occurred. Please try again later.')

    else:
        # Create instances of the login and registration forms
        login_form = UserLoginForm()
        register_form = UserRegisterForm()

    return render(request, 'taskmaster/auth.html', {'login_form': login_form, 'register_form': register_form, 'checked': checked})


def logout_view(request):
    """
    Logs out the current user and redirects to the authentication page.

    Parameters:
        request (HttpRequest): The request object.

    Returns:
        HttpResponseRedirect: Redirects the user to the authentication page after successful logout.

    """
    logout(request)  # Logs out the current user.
    # Displays a success message.
    messages.success(request, 'Logged out successfully.')
    return redirect('auth')  # Redirects to the authentication page.


def dailytask(request):
    """
    View function to display the daily task management page.

    If the user is not authenticated, it redirects to the authentication page.

    Parameters:
        request (HttpRequest): The HTTP request object.

    Returns:
        HttpResponse: The HTTP response object with the rendered dailytask.html template.
    """
    if not request.user.is_authenticated:
        # Redirects to the authentication page if the user is not authenticated.
        return redirect('auth')

    # Creates an instance of the TaskForm to be used in the template for adding new tasks.
    form = TaskForm()

    # Fetching daily tasks from the database for the current authenticated user, ordered by creation time.
    tasks = Task.objects.filter(
        daily=True, user=request.user).order_by('created_at')

    # Counting the number of completed tasks and total tasks for displaying statistics on the template.
    completed_tasks = tasks.filter(completed=True).count()
    total_tasks = tasks.count()

    # Rendering the dailytask.html template with the retrieved tasks, form, and task statistics as context data.
    return render(request, 'taskmaster/dailytask.html', {
        'tasks': tasks,
        'completed_tasks': completed_tasks,
        'total_tasks': total_tasks,
        'form': form,
    })


def weeklytask(request):
    """
    View function to handle displaying and managing weekly tasks.

    This view allows authenticated users to view their weekly tasks and manage them.
    Users can view, add, and mark tasks as completed for the current week.

    Args:
        request (HttpRequest): The HTTP request sent by the user.

    Returns:
        HttpResponse: The HTTP response containing the rendered weeklytask.html template
                      along with the context data (tasks, completed_tasks, total_tasks, form).
    """

    # Check if the user is authenticated. If not, redirect to the authentication page.
    if not request.user.is_authenticated:
        return redirect('auth')

    # Initialize an empty form to add new tasks.
    form = TaskForm()

    # Fetch all weekly tasks from the database for the current user, ordered by creation time.
    tasks = Task.objects.filter(
        weekly=True, user=request.user).order_by('created_at')

    # Count the number of completed weekly tasks.
    completed_tasks = tasks.filter(completed=True).count()

    # Count the total number of weekly tasks.
    total_tasks = tasks.count()

    # Render the weeklytask.html template with the context data (tasks, completed_tasks, total_tasks, form).
    return render(request, 'taskmaster/weeklytask.html', {
        'tasks': tasks,
        'completed_tasks': completed_tasks,
        'total_tasks': total_tasks,
        'form': form,
    })


def monthlytask(request):
    """
    View function for displaying monthly tasks.

    This view displays the list of monthly tasks for the authenticated user,
    along with the count of completed tasks and the total number of tasks.

    Parameters:
        request (HttpRequest): The HTTP request object.

    Returns:
        HttpResponse: The HTTP response with the rendered template.

    Redirects:
        - If the user is not authenticated, it redirects to the 'auth' URL.

    Template Context:
        - 'tasks': QuerySet of monthly tasks filtered by the authenticated user, ordered by 'created_at'.
        - 'completed_tasks': The count of completed monthly tasks.
        - 'total_tasks': The total count of monthly tasks.
        - 'form': An instance of the TaskForm to allow adding new tasks.
    """
    if not request.user.is_authenticated:
        # Redirect to the authentication page if the user is not logged in.
        return redirect('auth')

    # Create an instance of the TaskForm to display on the template.
    form = TaskForm()

    # Fetching monthly tasks from the database for the authenticated user, ordered by 'created_at'.
    tasks = Task.objects.filter(
        monthly=True, user=request.user).order_by('created_at')

    # Count the number of completed monthly tasks.
    completed_tasks = tasks.filter(completed=True).count()

    # Calculate the total number of monthly tasks.
    total_tasks = tasks.count()

    # Render the 'monthlytask.html' template with the task data and the TaskForm instance.
    return render(request, 'taskmaster/monthlytask.html', {'tasks': tasks, 'completed_tasks': completed_tasks, 'total_tasks': total_tasks, 'form': form})


@login_required
def add_task(request):
    """
    View function to handle adding a new task.

    This view handles the form submission to add a new task for the authenticated user.
    Depending on the task's frequency (daily, weekly, or monthly), the user will be redirected
    to the appropriate page after successful task addition.

    Parameters:
        request (HttpRequest): The HTTP request object.

    Returns:
        HttpResponse: The HTTP response, redirecting the user to the appropriate task page.
    """
    if request.method == 'POST' and request.user.is_authenticated:
        # Create a TaskForm instance and bind it to the POST data
        form = TaskForm(request.POST)

        # Check if the form data is valid
        if form.is_valid():
            # Create a new task object but do not save it to the database yet
            task = form.save(commit=False)

            # Associate the task with the current authenticated user
            task.user = request.user

            # Save the task to the database
            task.save()

            # Display a success message using Django's messages framework
            messages.success(
                request, f'Task <b>{task.title}</b> has been added')

            # Redirect the user to the appropriate task page based on the task's frequency
            if task.daily:
                return redirect('dailytask')
            elif task.weekly:
                return redirect('weeklytask')
            elif task.monthly:
                return redirect('monthlytask')
        else:
            # If the form data is invalid, display an error message
            messages.error(request, "An error occurred. Please try again")

            # Redirect the user to the appropriate task page based on the form data
            if form.cleaned_data.get('daily'):
                return redirect('dailytask')
            elif form.cleaned_data.get('weekly'):
                return redirect('weeklytask')
            elif form.cleaned_data.get('monthly'):
                return redirect('monthlytask')
            else:
                return redirect('index')


@login_required
def edit_task(request, task_id):
    """
    Edit an existing task with the given task_id belonging to the current user.

    If the HTTP method is POST and the form data is valid, update the task's information in the database.
    If the task is marked as daily, redirect to the daily task page after editing.
    If the task is marked as weekly, redirect to the weekly task page after editing.
    If the task is marked as monthly, redirect to the monthly task page after editing.
    If there is an error or the form data is invalid, display an error message and redirect to the appropriate task page.

    Args:
        request (HttpRequest): The HTTP request object containing the form data and user information.
        task_id (int): The ID of the task to be edited.

    Returns:
        HttpResponse: A redirect response to the appropriate task page or an error page if there is an error.
    """
    if request.method == 'POST':
        # Get the task object belonging to the current user with the given task_id
        task = Task.objects.get(id=task_id, user=request.user)
        form = TaskForm(request.POST, instance=task)
        if form.is_valid():
            # Save the updated task information to the database
            task = form.save()
            messages.success(
                request, f'Task <b>{task.title}</b> has been edited')
            # Redirect based on the task's frequency (daily, weekly, monthly)
            if task.daily:
                return redirect('dailytask')
            elif task.weekly:
                return redirect('weeklytask')
            elif task.monthly:
                return redirect('monthlytask')
        else:
            messages.error(request, f'An error occurred. Please try again.')
            # Redirect based on the task's frequency in the form data (daily, weekly, monthly)
            if form.cleaned_data.get('daily'):
                return redirect('dailytask')
            elif form.cleaned_data.get('weekly'):
                return redirect('weeklytask')
            elif form.cleaned_data.get('monthly'):
                return redirect('monthlytask')
            else:
                return redirect('index')


@login_required
def delete_task(request, task_id):
    """
    View function to delete a task.

    Args:
        request (HttpRequest): The HTTP request object.
        task_id (int): The ID of the task to be deleted.

    Returns:
        HttpResponse: Redirects to the appropriate task list page after successful deletion.
    """
    if request.method == 'POST':
        # Get the task object or return a 404 error if the task doesn't exist or doesn't belong to the user.
        task = get_object_or_404(Task, id=task_id, user=request.user)

        # Delete the task and display a success message.
        if task:
            task_title = task.title
            task.delete()
            messages.success(
                request, f'Task <b>{task_title}</b> has been deleted')

        # Redirect the user to the appropriate task list page based on the task type.
        if task.daily:
            return redirect('dailytask')
        elif task.weekly:
            return redirect('weeklytask')
        elif task.monthly:
            return redirect('monthlytask')


@login_required
def mark_task_complete(request, task_id):
    """
    View function to mark a task as complete or incomplete based on the request.

    Args:
        request (HttpRequest): The HTTP request object.
        task_id (int): The ID of the task to mark as complete or incomplete.

    Returns:
        HttpResponseRedirect: Redirects to the appropriate task view page (daily, weekly, monthly).
    """
    if request.method == 'POST':
        # Get the task object or return a 404 error if not found or not owned by the current user.
        task = get_object_or_404(Task, id=task_id, user=request.user)

        # Toggle the 'completed' status of the task.
        task.completed = not task.completed
        task.save()

        # Show a success message if the task is completed.
        if task.completed:
            messages.success(
                request, f'Task "{task.title}" has been completed.')

        # Redirect to the appropriate task view page based on the task's frequency.
        if task.daily:
            return redirect('dailytask')
        elif task.weekly:
            return redirect('weeklytask')
        elif task.monthly:
            return redirect('monthlytask')


def export_task_to_excel(request):
    """
    View function to export user's task history to an Excel file.

    Args:
        request: The HTTP request object.

    Returns:
        HttpResponse: A response containing the Excel file as an attachment.
    """

    if not request.user.is_authenticated:
        return redirect('auth')

    # Create a new Excel workbook
    workbook = Workbook()

    # Write headers for each table section
    task_types = ["Daily", "Weekly", "Monthly"]

    # Function to write section header
    def write_section_header(section_name):
        """
        Write the section header for each table.

        Args:
            section_name (str): The name of the task type section.

        Returns:
            None
        """
        nonlocal row_index
        sheet.merge_cells(start_row=row_index, start_column=1,
                          end_row=row_index, end_column=len(headers))
        merged_cell = sheet.cell(row=row_index, column=1, value=section_name)
        merged_cell.font = Font(bold=True)
        merged_cell.alignment = Alignment(horizontal='center')
        merged_cell.fill = PatternFill(
            start_color='999999', end_color='999999', fill_type='solid')

        for col_index, header in enumerate(headers, start=1):
            sheet.cell(row=row_index + 1, column=col_index,
                       value=header).font = Font(bold=True)
            sheet.cell(
                row=row_index + 1, column=col_index).alignment = Alignment(horizontal='center')
            sheet.cell(row=row_index + 1, column=col_index).fill = PatternFill(
                start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
            col_letter = get_column_letter(col_index)
            # Adjust width based on header length
            column_width = len(header) + 2
            sheet.column_dimensions[col_letter].width = column_width
        row_index += 2

    for task_type in task_types:
        # Create a new sheet for each task type
        sheet = workbook.create_sheet(title=task_type)

        # Write headers for each table section
        headers = ["Date", "Title", "Description",
                   "Execution_time", "Completed"]

        row_index = 1

        write_section_header(task_type)

        # Fetch and populate tasks for each type
        tasks = TaskHistory.objects.filter(
            task_type=task_type.lower(), user=request.user)
        for task in tasks:
            execution_time = task.execution_time.strftime('%H:%M')
            if task.task_type == 'daily':
                data = [task.date.strftime(
                    '%d-%m-%Y'), task.title, task.description, execution_time, 'yes' if task.completed else 'no']
            elif task.task_type == 'weekly':
                data = [task.date.strftime('%d-%m-%Y'), task.title, task.description,
                        f'{task.execution_day}, {execution_time}', 'yes' if task.completed else 'no']
            else:
                data = [task.date.strftime('%d-%m-%Y'), task.title, task.description,
                        f'Day {task.execution_date}, {execution_time}', 'yes' if task.completed else 'no']

            for col_index, value in enumerate(data, start=1):
                sheet.cell(row=row_index, column=col_index, value=value)
                col_letter = get_column_letter(col_index)
                # Adjust width based on content length
                column_width = len(str(value)) + 2
                if sheet.column_dimensions[col_letter].width < column_width:
                    sheet.column_dimensions[col_letter].width = column_width if column_width < 30 else 30
            row_index += 1

        # Add a border to the table
        table_start_row = row_index - len(tasks) - 2
        table_end_row = row_index - 1
        for col_index in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_index)
            side = Side(border_style='thin', color='000000')
            border = Border(top=side, bottom=side, left=side, right=side)
            for row in range(table_start_row, table_end_row + 1):
                sheet.cell(row=row, column=col_index).border = border

        row_index += 1

    # Remove the default 'Sheet' that is automatically created when the workbook is initialized
    workbook.remove(workbook['Sheet'])

    # Create the response with the Excel file
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f"attachment; filename=Task_Master_Data_{request.user.username}.xlsx"
    workbook.save(response)

    return response
